#!/usr/bin/env python3
"""
Lokale aggregatie -- rekent het agentic-team-metricsbestand uit een lokale
werkmap (agentic-team.xlsx of data/*.json), zonder ooit een rij door het
model te halen.

GEGENEREERD BESTAND -- dit bestand wordt door installer/build_plugin.py in
agent-architecture uit core/agents.json samengesteld en als
scripts/lokale_aggregatie.py in de plugin meegeleverd, zodat het op de
machine van elke klant beschikbaar is. Niet handmatig bewerken op een
klantmachine -- pas core/agents.json of build_plugin.py aan en genereer
opnieuw (deterministisch: dezelfde registry geeft altijd byte-identieke
uitvoer, dus alleen een echte registrywijziging verandert dit bestand).

  registryVersion : 1.37.0
  registry updated: 2026-08-25

Contract: exact dezelfde nieteentien sleutels als de Notion-route levert
(zie core/base/orchestrator/prompt.md, fase dagelijkse-metrics, en het
canonieke voorbeeld in agentic-team-dashboard/testdata/notion-metrics/
metrics.json). METRICS_VERSION hieronder moet gelijk blijven aan
METRICS_VERSION in agentic-team-dashboard/src/metrics.js.

Gebruik:
    python3 lokale_aggregatie.py [werkmap] [--vandaag JJJJ-MM-DD]
        [--weken N] [--minuten-per-actie N] [--geen-schrijven]

Leest ofwel <werkmap>/agentic-team.xlsx, ofwel <werkmap>/data/*.json.
Schrijft <werkmap>/agentic-team-metrics.json en drukt dezelfde JSON af op
stdout. Statusregels gaan naar stderr, zodat stdout altijd zuivere JSON is.

Xlsx vraagt het pakket openpyxl (niet in de standaardbibliotheek). Ontbreekt
dat, dan meldt dit script de installatieopdracht en stopt -- er wordt nooit
zelf aan xlsx-byte-parsing gedaan.
"""

import argparse
import json
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

METRICS_VERSION = 1  # moet gelijk blijven aan METRICS_VERSION in agentic-team-dashboard/src/metrics.js

# ── Schema (uit core/agents.json -- door build_plugin.py ingevuld) ─────────
SCHEMA = {
  "agents": [
    {
      "slug": "orchestrator",
      "displayName": "Coördinator",
      "emoji": "🔮",
      "module": "core"
    },
    {
      "slug": "management-assistent",
      "displayName": "Management Assistent",
      "emoji": "📌",
      "module": "core"
    },
    {
      "slug": "quality-control",
      "displayName": "Quality Control",
      "emoji": "🛡️",
      "module": "core"
    },
    {
      "slug": "gids",
      "displayName": "Gids",
      "emoji": "🧭",
      "module": "core"
    },
    {
      "slug": "ceo-agent",
      "displayName": "CEO Agent",
      "emoji": "🌟",
      "module": "strategy"
    },
    {
      "slug": "coo-agent",
      "displayName": "COO Agent",
      "emoji": "🏛️",
      "module": "backoffice"
    },
    {
      "slug": "marktmaker",
      "displayName": "Marktmaker",
      "emoji": "🎯",
      "module": "growth"
    },
    {
      "slug": "researcher",
      "displayName": "Researcher",
      "emoji": "🔍",
      "module": "growth"
    },
    {
      "slug": "pipeline-manager",
      "displayName": "Pipeline Manager",
      "emoji": "📊",
      "module": "growth"
    },
    {
      "slug": "product-designer",
      "displayName": "Product Designer",
      "emoji": "🧪",
      "module": "strategy"
    },
    {
      "slug": "outreach-specialist",
      "displayName": "Outreach Specialist",
      "emoji": "📨",
      "module": "sales"
    },
    {
      "slug": "dealmaker",
      "displayName": "Dealmaker",
      "emoji": "🤝",
      "module": "sales"
    },
    {
      "slug": "content-strateeg",
      "displayName": "Content Strateeg",
      "emoji": "✍️",
      "module": "visibility"
    },
    {
      "slug": "de-stem",
      "displayName": "De Stem",
      "emoji": "🎙️",
      "module": "visibility"
    },
    {
      "slug": "delivery-architect",
      "displayName": "Delivery Architect",
      "emoji": "🎒",
      "module": "delivery"
    },
    {
      "slug": "controller",
      "displayName": "Controller",
      "emoji": "📊",
      "module": "backoffice"
    },
    {
      "slug": "jurist",
      "displayName": "Jurist",
      "emoji": "⚖️",
      "module": "backoffice"
    },
    {
      "slug": "administratie",
      "displayName": "Administratie",
      "emoji": "📋",
      "module": "backoffice"
    },
    {
      "slug": "seo-geo-specialist",
      "displayName": "SEO/GEO Specialist",
      "emoji": "🧲",
      "module": "visibility"
    },
    {
      "slug": "customer-success-manager",
      "displayName": "Customer Success Manager",
      "emoji": "💚",
      "module": "delivery"
    }
  ],
  "datadomeinen": {
    "organisaties": {
      "naam": "Organisaties",
      "module": "core",
      "velden": [
        {
          "naam": "Naam"
        },
        {
          "naam": "Fase"
        },
        {
          "naam": "Segment"
        },
        {
          "naam": "Omvang FTE"
        },
        {
          "naam": "Lead Bron"
        },
        {
          "naam": "Signaal"
        },
        {
          "naam": "Signaal Datum"
        },
        {
          "naam": "Website"
        },
        {
          "naam": "LinkedIn"
        },
        {
          "naam": "Vestigingsplaats"
        },
        {
          "naam": "Eigenaar"
        },
        {
          "naam": "KvK-nummer"
        },
        {
          "naam": "Notities"
        }
      ]
    },
    "contactpersonen": {
      "naam": "Contactpersonen",
      "module": "core",
      "velden": [
        {
          "naam": "Naam"
        },
        {
          "naam": "Functie"
        },
        {
          "naam": "E-mail"
        },
        {
          "naam": "LinkedIn URL"
        },
        {
          "naam": "Warmte"
        },
        {
          "naam": "Outreach Status"
        },
        {
          "naam": "Rol in besluitvorming"
        },
        {
          "naam": "Laatste Contact"
        },
        {
          "naam": "Notities"
        }
      ]
    },
    "interacties": {
      "naam": "Interacties",
      "module": "sales",
      "velden": [
        {
          "naam": "Onderwerp"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Richting"
        },
        {
          "naam": "Datum"
        },
        {
          "naam": "Samenvatting"
        },
        {
          "naam": "Sentiment"
        },
        {
          "naam": "Volgende Actie"
        },
        {
          "naam": "Actie Deadline"
        },
        {
          "naam": "Bericht Tekst"
        },
        {
          "naam": "Organisatie"
        },
        {
          "naam": "Deal (link)"
        },
        {
          "naam": "Actie doorgezet"
        }
      ]
    },
    "sales_funnel": {
      "naam": "Sales Funnel",
      "module": "core",
      "velden": [
        {
          "naam": "Deal Naam"
        },
        {
          "naam": "Fase"
        },
        {
          "naam": "Opvolg Status"
        },
        {
          "naam": "Verwachte Omzet"
        },
        {
          "naam": "Probability"
        },
        {
          "naam": "Verwachte Sluitdatum"
        },
        {
          "naam": "Volgende Actie"
        },
        {
          "naam": "Volgende Actie Deadline"
        },
        {
          "naam": "Laatste Contact"
        },
        {
          "naam": "Lead Bron"
        },
        {
          "naam": "Eigenaar"
        },
        {
          "naam": "Notities"
        }
      ]
    },
    "projecten": {
      "naam": "Projecten",
      "module": "delivery",
      "velden": [
        {
          "naam": "Projectnaam"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Startdatum"
        },
        {
          "naam": "Einddatum"
        },
        {
          "naam": "Contractwaarde"
        },
        {
          "naam": "Gefactureerd"
        },
        {
          "naam": "Eigenaar"
        },
        {
          "naam": "Gedeelde Pagina"
        },
        {
          "naam": "Notities"
        }
      ]
    },
    "offertes": {
      "naam": "Offertes",
      "module": "sales",
      "velden": [
        {
          "naam": "Offertenummer"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Bedrag excl. BTW"
        },
        {
          "naam": "BTW bedrag"
        },
        {
          "naam": "Geldig tot"
        },
        {
          "naam": "Verstuurd op"
        },
        {
          "naam": "Getekend op"
        },
        {
          "naam": "Offerte URL"
        },
        {
          "naam": "Notities"
        }
      ]
    },
    "product_catalogus": {
      "naam": "Product Catalogus",
      "module": "strategy",
      "velden": [
        {
          "naam": "Productnaam"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Prijs"
        },
        {
          "naam": "Prijsmodel"
        },
        {
          "naam": "Plek in ladder"
        },
        {
          "naam": "Segment"
        },
        {
          "naam": "Beschrijving"
        },
        {
          "naam": "USP"
        },
        {
          "naam": "Notities"
        }
      ]
    },
    "content_kalender": {
      "naam": "Content Kalender",
      "module": "visibility",
      "velden": [
        {
          "naam": "Titel"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Segment"
        },
        {
          "naam": "Publicatiedatum"
        },
        {
          "naam": "Kanaal"
        },
        {
          "naam": "Hook A"
        },
        {
          "naam": "Hook B"
        },
        {
          "naam": "Concept Tekst"
        },
        {
          "naam": "Campagne"
        },
        {
          "naam": "Eigenaar"
        }
      ]
    },
    "lessen_inzichten": {
      "naam": "Lessen & Inzichten",
      "module": "core",
      "velden": [
        {
          "naam": "Les"
        },
        {
          "naam": "Agent"
        },
        {
          "naam": "Categorie"
        },
        {
          "naam": "Datum"
        },
        {
          "naam": "Actie"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Impact"
        },
        {
          "naam": "Verwerkt in prompt?"
        }
      ]
    },
    "bedrijfscontext": {
      "naam": "Bedrijfscontext",
      "module": "core",
      "velden": [
        {
          "naam": "Onderdeel"
        },
        {
          "naam": "Inhoud"
        },
        {
          "naam": "Versie"
        },
        {
          "naam": "Bijgewerkt"
        },
        {
          "naam": "Status"
        }
      ]
    },
    "logboek": {
      "naam": "Logboek",
      "module": "core",
      "velden": [
        {
          "naam": "Onderwerp"
        },
        {
          "naam": "Agent"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Datum"
        },
        {
          "naam": "Resultaat"
        },
        {
          "naam": "Link"
        },
        {
          "naam": "Vervolg"
        }
      ]
    },
    "dagverslagen": {
      "naam": "Dagverslagen",
      "module": "core",
      "velden": [
        {
          "naam": "Naam"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Dag"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Persoon"
        }
      ]
    },
    "delivery_rugzak": {
      "naam": "Delivery Rugzak",
      "module": "delivery",
      "velden": [
        {
          "naam": "Naam"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Beschrijving"
        },
        {
          "naam": "Wanneer inzetten"
        },
        {
          "naam": "Benodigdheden"
        },
        {
          "naam": "Segment"
        },
        {
          "naam": "Bron"
        }
      ]
    },
    "tijdregistratie": {
      "naam": "Tijdregistratie",
      "module": "backoffice",
      "velden": [
        {
          "naam": "Beschrijving"
        },
        {
          "naam": "Datum"
        },
        {
          "naam": "Uren"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Uurtarief"
        },
        {
          "naam": "Gefactureerd"
        },
        {
          "naam": "Persoon"
        }
      ]
    },
    "acties": {
      "naam": "Acties",
      "module": "core",
      "velden": [
        {
          "naam": "Actie"
        },
        {
          "naam": "Eigenaar"
        },
        {
          "naam": "Agent"
        },
        {
          "naam": "Deadline"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Type"
        },
        {
          "naam": "Prioriteit"
        },
        {
          "naam": "Toelichting"
        },
        {
          "naam": "Aangemaakt door"
        },
        {
          "naam": "Organisatie"
        },
        {
          "naam": "Bron (link)"
        },
        {
          "naam": "Afgerond door"
        },
        {
          "naam": "Afgerond op"
        },
        {
          "naam": "Gecorrigeerd"
        },
        {
          "naam": "Correctie"
        }
      ]
    },
    "ritmetaken": {
      "naam": "Ritmetaken",
      "module": "core",
      "velden": [
        {
          "naam": "Taak"
        },
        {
          "naam": "Agent"
        },
        {
          "naam": "Ritme"
        },
        {
          "naam": "Volgorde"
        },
        {
          "naam": "Laatst gedraaid"
        },
        {
          "naam": "Actief"
        },
        {
          "naam": "Bron-template"
        },
        {
          "naam": "Instructie"
        }
      ]
    },
    "klantsucces": {
      "naam": "Klantsucces",
      "module": "delivery",
      "velden": [
        {
          "naam": "Klantnaam"
        },
        {
          "naam": "Fase"
        },
        {
          "naam": "Health"
        },
        {
          "naam": "Verlengdatum"
        },
        {
          "naam": "Laatste check"
        },
        {
          "naam": "Signalen"
        },
        {
          "naam": "Eigenaar"
        }
      ]
    },
    "productbacklog": {
      "naam": "Productbacklog",
      "module": "core",
      "velden": [
        {
          "naam": "Item"
        },
        {
          "naam": "Herkomst"
        },
        {
          "naam": "Eigenaar"
        },
        {
          "naam": "Status"
        },
        {
          "naam": "Prioriteit"
        },
        {
          "naam": "Besluit"
        }
      ]
    },
    "dashboard_metrics": {
      "naam": "Dashboardmetrics",
      "module": "core",
      "velden": [
        {
          "naam": "Titel"
        },
        {
          "naam": "Inhoud"
        }
      ]
    },
    "bronkoppeling": {
      "naam": "Bronkoppeling",
      "module": "core",
      "velden": [
        {
          "naam": "Titel"
        },
        {
          "naam": "Systeem"
        },
        {
          "naam": "Verwijzing"
        },
        {
          "naam": "Veldvertaling"
        },
        {
          "naam": "Laatst_geverifieerd"
        },
        {
          "naam": "Notitie"
        }
      ]
    },
    "teamfeed": {
      "naam": "Teamfeed",
      "module": "core",
      "velden": [
        {
          "naam": "Actie"
        },
        {
          "naam": "Agent"
        },
        {
          "naam": "Soort"
        },
        {
          "naam": "Bericht"
        },
        {
          "naam": "Link"
        }
      ]
    }
  }
}

RITME_BRONNEN = ["dagverslagen", "lessen_inzichten", "interacties", "content_kalender"]
RITME_DATUMVELD = {
    "dagverslagen": "Dag",
    "lessen_inzichten": "Datum",
    "interacties": "Datum",
    "content_kalender": "Publicatiedatum",
}

# Zelfde tabel als agentic-team-dashboard/src/schema-helpers.js -- oudere/
# alternatieve bestandsnamen die niet letterlijk de registry-domeinsleutel
# volgen. Puur voor herkenning bij het inlezen.
LEGACY_FILENAME_ALIASES = {
    "prospects": "organisaties",
    "lessen": "lessen_inzichten",
    "content-pipeline": "content_kalender",
    "content-kalender": "content_kalender",
    "sales-funnel": "sales_funnel",
}


# ── Kleine hulpfuncties (spiegelen schema-helpers.js / zones.js) ───────────
def norm_key(s):
    """Zelfde normalisatie als normKey() in schema-helpers.js: lowercase,
    diakrieten weg, alleen a-z0-9 over."""
    if s is None:
        return ""
    s = str(s).lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s)


def get_field(row, canonical_name):
    """Zoekt een veldwaarde op, eerst exact, dan genormaliseerd -- zelfde
    gedrag als getField() in schema-helpers.js."""
    if row is None:
        return None
    if canonical_name in row:
        return row[canonical_name]
    target = norm_key(canonical_name)
    for k, v in row.items():
        if norm_key(k) == target:
            return v
    return None


def parse_date_field(v):
    """Parseert een datumveld naar een date-object. Accepteert ISO-strings
    (met of zonder tijdcomponent) en Python date/datetime-objecten (zoals
    openpyxl die voor datumcellen kan teruggeven). Geen waarde of onherkenbaar
    -> None (zelfde als parseDateField() dat NaN-datums als null behandelt)."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # Alleen het datumdeel gebruiken (eerste 10 tekens dekken JJJJ-MM-DD;
    # bij een volledige ISO-tijdstempel proberen we die eerst).
    try:
        if len(s) > 10 and ("T" in s or " " in s):
            return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def days_between(a, b):
    """a - b in dagen -- zelfde teken-conventie als daysBetween() in
    zones.js: daysBetween(today, dt) is positief als dt in het verleden
    ligt."""
    return (a - b).days


def build_agent_lookup(schema):
    lookup = {}
    for agent in schema["agents"]:
        variants = [
            agent["slug"],
            agent["displayName"],
            f"{agent.get('emoji', '')} {agent['displayName']}",
            agent["displayName"].replace(" ", "-"),
        ]
        for v in variants:
            lookup[norm_key(v)] = agent["slug"]
    return lookup


def match_agent_value(raw, lookup):
    if not raw:
        return None
    return lookup.get(norm_key(raw))


def domain_key_from_filename(filename, schema):
    base = re.sub(r"\.[a-zA-Z0-9]+$", "", filename)
    kebab = base.lower()
    if kebab in LEGACY_FILENAME_ALIASES:
        return LEGACY_FILENAME_ALIASES[kebab]
    for key in schema["datadomeinen"].keys():
        if norm_key(key) == norm_key(base):
            return key
    return None


def domain_key_from_sheet_name(sheet_name, schema):
    for key, domein in schema["datadomeinen"].items():
        if domein.get("naam") == sheet_name:
            return key
    return None


def domain_key_from_headers(headers, schema):
    norm_headers = {norm_key(h) for h in headers}
    best, best_score = None, 0
    for key, domein in schema["datadomeinen"].items():
        velden = [norm_key(v["naam"]) for v in domein.get("velden", [])]
        if not velden:
            continue
        overlap = sum(1 for v in velden if v in norm_headers)
        score = overlap / len(velden)
        if score > best_score:
            best_score, best = score, key
    return best if best_score >= 0.5 else None


# ── Bundel inlezen ──────────────────────────────────────────────────────
class Bundel:
    def __init__(self, bron_label):
        self.bron_label = bron_label
        self.domains = {}  # key -> {"rows": [...], "mtime": date}
        self.bedrijfscontext = None  # None of dict met "staleAt": date
        self.waarschuwingen = []


def _mtime_date(path):
    return datetime.fromtimestamp(path.stat().st_mtime).date()


def load_json_bundle(data_dir, schema):
    bundel = Bundel("data/-map (lokale bestanden)")
    for f in sorted(data_dir.glob("*.json")):
        try:
            parsed = json.loads(f.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            bundel.waarschuwingen.append(f"{f.name} is geen geldige JSON -- genegeerd ({e}).")
            continue

        if f.stem.lower() == "bedrijfscontext":
            laatst = parsed.get("Laatst_bijgewerkt") or parsed.get("laatst_bijgewerkt")
            geexporteerd = parsed.get("_geexporteerd_op")
            if laatst:
                stale = parse_date_field(laatst)
            elif geexporteerd:
                stale = parse_date_field(geexporteerd)
            else:
                stale = _mtime_date(f)
            bedrijfscontext = dict(parsed)
            bedrijfscontext["staleAt"] = stale
            bundel.bedrijfscontext = bedrijfscontext
            continue

        domain_key = domain_key_from_filename(f.name, schema)
        if not domain_key:
            items = parsed if isinstance(parsed, list) else parsed.get("items")
            if items:
                domain_key = domain_key_from_headers(list(items[0].keys()), schema)
        if not domain_key:
            bundel.waarschuwingen.append(f"{f.name} kon niet aan een bekend datadomein gekoppeld worden -- genegeerd.")
            continue

        items = parsed if isinstance(parsed, list) else parsed.get("items", [])
        bundel.domains[domain_key] = {"rows": items, "mtime": _mtime_date(f)}
    return bundel


def load_excel_bundle(xlsx_path, schema):
    try:
        import openpyxl
    except ImportError:
        print(
            "FOUT: het pakket 'openpyxl' is niet geinstalleerd -- dit is nodig om "
            f"{xlsx_path.name} te lezen. Installeer het met:\n"
            "    python3 -m pip install openpyxl\n"
            "en draai dit script daarna opnieuw. Er is niet geprobeerd het "
            "werkboek zonder openpyxl te lezen.",
            file=sys.stderr,
        )
        sys.exit(1)

    bundel = Bundel(f"Excel-werkboek ({xlsx_path.name})")
    file_mtime = _mtime_date(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name == "_schema":
            continue
        ws = wb[sheet_name]
        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            continue
        headers = [str(h) if h is not None else "" for h in rows_iter[0]]

        if sheet_name.lower() == "bedrijfscontext":
            obj = {}
            for r in rows_iter[1:]:
                if r and r[0]:
                    obj[str(r[0])] = r[1] if len(r) > 1 else None
            laatst = obj.get("Laatst_bijgewerkt")
            obj["staleAt"] = parse_date_field(laatst) if laatst else file_mtime
            bundel.bedrijfscontext = obj
            continue

        domain_key = domain_key_from_sheet_name(sheet_name, schema)
        if not domain_key:
            domain_key = domain_key_from_headers(headers, schema)
        if not domain_key:
            bundel.waarschuwingen.append(f'Tabblad "{sheet_name}" kon niet aan een bekend datadomein gekoppeld worden -- genegeerd.')
            continue

        data_rows = [r for r in rows_iter[1:] if any(c not in (None, "") for c in r)]
        obj_rows = []
        for r in data_rows:
            obj = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                obj[h] = r[i] if i < len(r) else None
            obj_rows.append(obj)

        bundel.domains[domain_key] = {"rows": obj_rows, "mtime": file_mtime}

    return bundel


# ── Aggregaties (spiegelen agentic-team-dashboard/src/zones.js) ────────────
def compute_domeinen_block(bundel):
    return {
        key: {"rijen": len(dom["rows"]), "laatst_bijgewerkt": dom["mtime"].isoformat()}
        for key, dom in bundel.domains.items()
    }


def compute_weekreeks_block(bundel, today, weeks):
    period_start = today - timedelta(days=weeks * 7)
    aanwezig = [b for b in RITME_BRONNEN if b in bundel.domains]
    buckets = [{k: 0 for k in RITME_BRONNEN} for _ in range(weeks)]

    for bron in RITME_BRONNEN:
        dom = bundel.domains.get(bron)
        if not dom:
            continue
        veld = RITME_DATUMVELD[bron]
        for row in dom["rows"]:
            d = parse_date_field(get_field(row, veld))
            if not d or d > today or d < period_start:
                continue
            idx = (d - period_start).days // 7
            if idx >= weeks:
                idx = weeks - 1
            buckets[idx][bron] += 1

    out_buckets = []
    for i, waarden in enumerate(buckets):
        week_start = period_start + timedelta(days=i * 7)
        out_buckets.append(
            {
                "week_start": week_start.isoformat(),
                "label": week_start.strftime("%d-%m"),
                "waarden": waarden,
                "totaal": sum(waarden.values()),
            }
        )
    return {"bronnen": aanwezig, "buckets": out_buckets}


def compute_agents_block(bundel, agent_lookup, schema, today, period_days):
    acties_dom = bundel.domains.get("acties")
    lessen_dom = bundel.domains.get("lessen_inzichten")
    if not acties_dom and not lessen_dom:
        return None  # bron ontbreekt -- blok weglaten, nooit nullen invullen

    slugs = [a["slug"] for a in schema["agents"]]
    traces = {slug: {"aantal_periode": 0, "aantal_totaal": 0, "laatst": None} for slug in slugs}
    veld_aanwezig = {"v": False}

    def registreer(slug_raw, datum_raw):
        if slug_raw is not None and str(slug_raw).strip() != "":
            veld_aanwezig["v"] = True
        slug = match_agent_value(slug_raw, agent_lookup)
        if not slug or slug not in traces:
            return
        dt = parse_date_field(datum_raw)
        t = traces[slug]
        t["aantal_totaal"] += 1
        if dt and (t["laatst"] is None or dt > t["laatst"]):
            t["laatst"] = dt
        if dt:
            diff = days_between(today, dt)
            if 0 <= diff <= period_days:
                t["aantal_periode"] += 1

    if acties_dom:
        for row in acties_dom["rows"]:
            registreer(get_field(row, "Agent"), get_field(row, "Deadline"))
    if lessen_dom:
        for row in lessen_dom["rows"]:
            registreer(get_field(row, "Agent"), get_field(row, "Datum"))

    per_agent = {}
    for slug, t in traces.items():
        if t["aantal_totaal"] == 0:
            continue
        entry = {"aantal_periode": t["aantal_periode"], "aantal_totaal": t["aantal_totaal"]}
        if t["laatst"]:
            entry["laatst"] = t["laatst"].isoformat()
        per_agent[slug] = entry

    return {"veld_aanwezig": veld_aanwezig["v"], "per_agent": per_agent}


def compute_acties_block(bundel, today):
    dom = bundel.domains.get("acties")
    if not dom:
        return None
    rows = dom["rows"]
    totaal = len(rows)
    afgerond = sum(1 for r in rows if get_field(r, "Status") == "Klaar")
    verstreken_rows = []
    for r in rows:
        dt = parse_date_field(get_field(r, "Deadline"))
        if dt and days_between(today, dt) > 0:  # deadline ligt voor vandaag
            verstreken_rows.append(r)
    klaar_verstreken = sum(1 for r in verstreken_rows if get_field(r, "Status") == "Klaar")
    return {
        "totaal": totaal,
        "afgerond": afgerond,
        "verstreken": len(verstreken_rows),
        "klaar_verstreken": klaar_verstreken,
        "opmerking": "Totaal = alle acties in de bundel (geen aanmaakdatum beschikbaar in dit domein om op periode te filteren).",
    }


def is_checkbox_true(v):
    """Checkbox-waarde uit werkboek/JSON: True, 1, "true", "ja", "x", "__YES__"
    (Notion-SQL) tellen als aangevinkt -- zelfde regel als het dashboard."""
    if v is True or v == 1:
        return True
    return str(v or "").strip().lower() in ("true", "ja", "x", "__yes__", "1")


def compute_correctievrij_block(bundel, today):
    """i25 -- correctievrij-percentage (f9-succesmaat, f19-gate). Alleen
    tellingen; percentage en gate rekent het dashboard zelf uit, zodat elke
    route identiek rekent. Autonoom afgerond = Afgerond door gevuld en
    Afgerond op binnen het venster; gecorrigeerd = Gecorrigeerd aangevinkt of
    Status != Klaar (heropend). Weken = kalenderweken (maandag), laatste 5."""
    dom = bundel.domains.get("acties")
    if not dom:
        return None
    venster = 28
    deze_maandag = today - timedelta(days=today.weekday())
    week_van = deze_maandag - timedelta(days=28)
    weken = {}
    tot = {"autonoom_afgerond": 0, "gecorrigeerd": 0, "heropend": 0}
    for r in dom["rows"]:
        if not str(get_field(r, "Afgerond door") or "").strip():
            continue
        dt = parse_date_field(get_field(r, "Afgerond op"))
        if not dt or dt > today:
            continue
        heropend = get_field(r, "Status") != "Klaar"
        gecorrigeerd = heropend or is_checkbox_true(get_field(r, "Gecorrigeerd"))
        if days_between(today, dt) < venster:
            tot["autonoom_afgerond"] += 1
            tot["gecorrigeerd"] += 1 if gecorrigeerd else 0
            tot["heropend"] += 1 if heropend else 0
        if dt >= week_van:
            ws = (dt - timedelta(days=dt.weekday())).isoformat()
            w = weken.setdefault(ws, {"week_start": ws, "autonoom_afgerond": 0, "gecorrigeerd": 0})
            w["autonoom_afgerond"] += 1
            w["gecorrigeerd"] += 1 if gecorrigeerd else 0
    return {
        "venster_dagen": venster,
        "drempel_pct": 80,
        **tot,
        "weken": [weken[k] for k in sorted(weken)],
        "opmerking": "Tellingen op de velden Afgerond door/Afgerond op/Gecorrigeerd/Status van Acties; percentage en gate rekent het dashboard.",
    }


def compute_sales_funnel_block(bundel):
    dom = bundel.domains.get("sales_funnel")
    if not dom:
        return None
    rows = dom["rows"]
    per_fase = {}
    omzet = 0.0
    for r in rows:
        fase = get_field(r, "Fase") or "Onbekend"
        per_fase[fase] = per_fase.get(fase, 0) + 1
        v = get_field(r, "Verwachte Omzet")
        try:
            omzet += float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            pass
    return {
        "per_fase": per_fase,
        "verwachte_omzet_totaal": omzet if omzet % 1 else int(omzet),
        "opmerking": "Huidige verdeling -- de bundel bevat geen wijzigingsgeschiedenis, dus geen trend van deals die van fase wisselden.",
    }


def compute_content_block(bundel, today, period_days):
    dom = bundel.domains.get("content_kalender")
    if not dom:
        return None
    rows = dom["rows"]
    gepubliceerd = sum(1 for r in rows if get_field(r, "Status") == "Gepubliceerd")
    gepland_in_periode = 0
    for r in rows:
        dt = parse_date_field(get_field(r, "Publicatiedatum"))
        if dt and abs(days_between(today, dt)) <= period_days and get_field(r, "Status") in ("Gepland", "Gepubliceerd"):
            gepland_in_periode += 1
    return {"gepubliceerd": gepubliceerd, "gepland_in_periode": gepland_in_periode, "totaal": len(rows)}


def compute_klantsucces_block(bundel):
    dom = bundel.domains.get("klantsucces")
    if not dom:
        return None
    rows = dom["rows"]
    return {"in_onboarding": sum(1 for r in rows if get_field(r, "Fase") == "Onboarding"), "totaal": len(rows)}


def compute_backlog_block(bundel):
    dom = bundel.domains.get("productbacklog")
    if not dom:
        return None
    rows = dom["rows"]
    besloten = sum(1 for r in rows if str(get_field(r, "Besluit") or "").strip() != "")
    done = sum(1 for r in rows if get_field(r, "Status") == "Done")
    return {"besloten": besloten, "done": done, "totaal": len(rows)}


def compute_lessen_block(bundel, today, period_days):
    dom = bundel.domains.get("lessen_inzichten")
    if dom is None:
        return None  # domein ontbreekt -> dashboard toont "aanwezig: false"
    rows = dom["rows"]
    if not rows:
        return {"totaal": 0}  # domein aanwezig maar leeg -> dashboard toont "leeg: true"
    per_categorie = {}
    open_n = 0
    in_periode = 0
    for r in rows:
        cat = get_field(r, "Categorie") or "Onbekend"
        per_categorie[cat] = per_categorie.get(cat, 0) + 1
        if get_field(r, "Status") == "Open":
            open_n += 1
        dt = parse_date_field(get_field(r, "Datum"))
        if dt and 0 <= days_between(today, dt) <= period_days:
            in_periode += 1
    return {"totaal": len(rows), "per_categorie": per_categorie, "open": open_n, "in_periode": in_periode}


def compute_bedrijfscontext_block(bundel):
    ctx = bundel.bedrijfscontext
    if ctx is None:
        return None
    bron = ctx.get("Bron") or ctx.get("bron")
    stale = ctx.get("staleAt")
    open_ph = ctx.get("Placeholders_open") or ctx.get("placeholders_open") or []
    kopie = ctx.get("Projectkennis_kopie_laatst_bijgewerkt") or ctx.get("projectkennis_kopie_laatst_bijgewerkt")
    return {
        "bron": bron,
        "laatst_bijgewerkt": stale.isoformat() if isinstance(stale, date) else stale,
        "placeholders_open": open_ph,
        "projectkennis_kopie_laatst_bijgewerkt": kopie,
    }


def compute_aandacht(bundel, agent_lookup, today):
    """Zelfde signalen als computeZone1() in agentic-team-dashboard/src/zones.js
    (QC-bevindingen, acties over de deadline, klantsucces-risico, stilstaande
    deals) -- als tellingen/labels, nooit als rijen. Gebruikt bewust de
    correcte "verleden"-richting (days_between(today, dt) > 0, zoals ook
    computeOpvolging() in dezelfde zones.js-module doet, met het commentaar
    "deadline ligt voor vandaag"). LET OP: computeZone1() zelf gebruikt op dit
    moment de omgekeerde richting (< 0) voor deze twee signalen -- dat lijkt
    een tekenfout in zones.js, zie het rapport bij deze wijziging. Dit script
    volgt bewust NIET die (vermoedelijk foutieve) richting."""
    items = []

    acties = bundel.domains.get("acties")
    if acties:
        rows = acties["rows"]
        qc_open = [
            r for r in rows
            if match_agent_value(get_field(r, "Agent"), agent_lookup) == "quality-control"
            and get_field(r, "Status") != "Klaar"
        ]
        if qc_open:
            items.append({"type": "qc", "ernst": "rood", "label": f"{len(qc_open)} QC-bevinding(en) die een menselijke blik vragen", "link": None})

        over_deadline = []
        for r in rows:
            dt = parse_date_field(get_field(r, "Deadline"))
            if dt and days_between(today, dt) > 0 and get_field(r, "Status") != "Klaar":
                over_deadline.append(r)
        if over_deadline:
            items.append({"type": "acties-deadline", "ernst": "rood", "label": f"{len(over_deadline)} actie(s) over de deadline", "link": None})

    ks = bundel.domains.get("klantsucces")
    if ks:
        risico = [r for r in ks["rows"] if get_field(r, "Health") in ("Oranje", "Rood")]
        if risico:
            ernst = "rood" if any(get_field(r, "Health") == "Rood" for r in risico) else "oranje"
            items.append({"type": "klantsucces", "ernst": ernst, "label": f"{len(risico)} klant(en) op oranje of rood in Klantsucces", "link": None})

    sf = bundel.domains.get("sales_funnel")
    if sf:
        stil = []
        for r in sf["rows"]:
            dt = parse_date_field(get_field(r, "Volgende Actie Deadline"))
            status = get_field(r, "Opvolg Status")
            if dt and days_between(today, dt) > 0 and status not in ("Gewonnen", "Verloren"):
                stil.append(r)
        if stil:
            items.append({"type": "deals-stil", "ernst": "oranje", "label": f"{len(stil)} deal(s) met een verlopen vervolgactie", "link": None})

    order = {"rood": 0, "grijs": 1, "oranje": 2, "groen": 3}
    items.sort(key=lambda it: order.get(it["ernst"], 9))
    return items[:5]


def build_metrics(bundel, schema, today, weeks, minuten_per_actie, door):
    agent_lookup = build_agent_lookup(schema)
    period_days = weeks * 7
    period_start = today - timedelta(days=period_days)

    metrics = {
        "type": "agentic-team-metrics",
        "versie": METRICS_VERSION,
        "bron_label": bundel.bron_label,
        "gegenereerd_op": datetime.now().astimezone().isoformat(timespec="seconds"),
        "door": door,
        "periode": {"van": period_start.isoformat(), "tot": today.isoformat(), "weken": weeks},
        "minuten_per_actie": minuten_per_actie,
        "domeinen": compute_domeinen_block(bundel),
        "weekreeks": compute_weekreeks_block(bundel, today, weeks),
    }

    agents_block = compute_agents_block(bundel, agent_lookup, schema, today, period_days)
    if agents_block is not None:
        metrics["agents"] = agents_block

    for key, value in (
        ("acties", compute_acties_block(bundel, today)),
        ("correctievrij", compute_correctievrij_block(bundel, today)),
        ("sales_funnel", compute_sales_funnel_block(bundel)),
        ("content", compute_content_block(bundel, today, period_days)),
        ("klantsucces", compute_klantsucces_block(bundel)),
        ("backlog", compute_backlog_block(bundel)),
        ("lessen", compute_lessen_block(bundel, today, period_days)),
        ("bedrijfscontext", compute_bedrijfscontext_block(bundel)),
    ):
        if value is not None:
            metrics[key] = value

    metrics["aandacht"] = compute_aandacht(bundel, agent_lookup, today)
    metrics["waarschuwingen"] = bundel.waarschuwingen

    return metrics


def main():
    parser = argparse.ArgumentParser(description="Rekent het agentic-team-metricsbestand uit een lokale werkmap (xlsx of data/-map).")
    parser.add_argument("werkmap", nargs="?", default=".", help="Pad naar de werkmap (standaard: huidige map)")
    parser.add_argument("--vandaag", default=None, help="Peildatum JJJJ-MM-DD (standaard: vandaag) -- vooral voor reproduceerbaar testen")
    parser.add_argument("--weken", type=int, default=12, help="Periode in weken voor de weekreeks/tellingen (standaard: 12)")
    parser.add_argument("--minuten-per-actie", type=int, default=25, dest="minuten_per_actie", help="Aanname achter de tijdwinst-schatting (standaard: 25)")
    parser.add_argument("--door", default="Coördinator (lokale aggregatie)", help='Wie dit schreef, voor het veld "door"')
    parser.add_argument("--geen-schrijven", action="store_true", help="Niet naar agentic-team-metrics.json schrijven, alleen naar stdout printen")
    args = parser.parse_args()

    werkmap = Path(args.werkmap).resolve()
    today = date.fromisoformat(args.vandaag) if args.vandaag else date.today()

    xlsx_path = werkmap / "agentic-team.xlsx"
    data_dir = werkmap / "data"

    if xlsx_path.exists() and data_dir.is_dir():
        print(
            f"Zowel {xlsx_path.name} als data/ gevonden in {werkmap} -- dit script gebruikt het werkboek. "
            "Verwijder een van beide om dubbelzinnigheid te voorkomen.",
            file=sys.stderr,
        )

    if xlsx_path.exists():
        bundel = load_excel_bundle(xlsx_path, SCHEMA)
    elif data_dir.is_dir():
        bundel = load_json_bundle(data_dir, SCHEMA)
    else:
        print(
            f"FOUT: geen {xlsx_path.name} of data/-map gevonden in {werkmap}. "
            "Dit script verwacht een van beide lokale routes.",
            file=sys.stderr,
        )
        sys.exit(1)

    metrics = build_metrics(bundel, SCHEMA, today, args.weken, args.minuten_per_actie, args.door)
    output = json.dumps(metrics, ensure_ascii=False, indent=2)

    if not args.geen_schrijven:
        out_path = werkmap / "agentic-team-metrics.json"
        tmp_path = out_path.with_suffix(".json.tmp")
        tmp_path.write_text(output + "\n", encoding="utf-8")
        tmp_path.replace(out_path)  # atomisch -- nooit een half bestand
        print(f"OK: {out_path} geschreven ({len(bundel.domains)} domeinen, periode {args.weken} weken).", file=sys.stderr)

    print(output)


if __name__ == "__main__":
    main()
